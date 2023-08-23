import openpyxl
import operator
import re
from openpyxl import workbook
from Version import *

def getMatrixSheet(matrixFile):
    workbook = openpyxl.load_workbook(matrixFile)
    matrixSheet = "Sheet1"

    for sheetName in workbook.sheetnames:
        if "Matrix" in sheetName:
            matrixSheet = sheetName
            break

    return workbook[matrixSheet]

def getMatrixNS(matrixFile):
    matrixSheet = getMatrixSheet(matrixFile)
    NS = '''\tNS_DESC_
\tCM_
\tBA_DEF_
\tBA_
\tVAL_
\tCAT_DEF_
\tCAT_
\tFILTER
\tBA_DEF_DEF_
\tEV_DATA_
\tENVVAR_DATA_
\tSGTYPE_
\tSGTYPE_VAL_
\tBA_DEF_SGTYPE_
\tBA_SGTYPE_
\tSIG_TYPE_REF_
\tVAL_TABLE_
\tSIG_GROUP_
\tSIG_VALTYPE_
\tSIGTYPE_VALTYPE_
\tBO_TX_BU_
\tBA_DEF_REL_
\tBA_REL_
\tBA_DEF_DEF_REL_
\tBU_SG_REL_
\tBU_EV_REL_
\tBU_BO_REL_
\tSG_MUL_VAL_'''
    return NS

def getMatrixBS(matrixFile):
    matrixSheet = getMatrixSheet(matrixFile)

    return ""

def getMatrixBUSInfo(matrixSheet):
    BUS = []
    BULoop = 0

    while(matrixSheet.cell(1, 31 + BULoop).value):
        BU = {}
        BU['name'] = matrixSheet.cell(1, 31 + BULoop).value
        BUS.append(BU)
        BULoop += 1

    return BUS

def getMatrixBUS(matrixFile):
    matrixSheet = getMatrixSheet(matrixFile)
    BUSName = ""

    for BUEle in getMatrixBUSInfo(matrixSheet):
        if not operator.eq(BUEle, getMatrixBUSInfo(matrixSheet)[0]):
            BUSName += " "
        BUSName += BUEle['name']

    return BUSName

def getAttrColumn(matrixSheet, item):
    col = 1

    while(matrixSheet.cell(1, col).value):
        if re.search(item, matrixSheet.cell(1, col).value, re.IGNORECASE):
            return col
        else:
            col += 1
    return 255

def getMsgInfo(matrixSheet, row):
    attrList = [
        {"item": "name",            "col": getAttrColumn(matrixSheet, "Msg Name")},
        {"item": "type",            "col": getAttrColumn(matrixSheet, "Msg Type")},
        {"item": "id",              "col": getAttrColumn(matrixSheet, "Msg ID")},
        {"item": "sendType",        "col": getAttrColumn(matrixSheet, "Msg Send Type")},
        {"item": "cycleTime",       "col": getAttrColumn(matrixSheet, "Msg Cycle Time")},
        {"item": "frameFormat",     "col": getAttrColumn(matrixSheet, "Frame Format")},
        {"item": "baudRateSwitch",  "col": getAttrColumn(matrixSheet, "BRS")},
        {"item": "len",             "col": getAttrColumn(matrixSheet, "Msg Length")},
        {"item": "desc",            "col": getAttrColumn(matrixSheet, "Signal Description")},
        {"item": "fastCycleTime",   "col": getAttrColumn(matrixSheet, "Msg Cycle Time Fast")},
        {"item": "numOfRepetition", "col": getAttrColumn(matrixSheet, "Msg Nr. Of Reption")},
        {"item": "delayTime",       "col": getAttrColumn(matrixSheet, "Msg Delay Time")},
    ]

    msg = {attrEle["item"] : matrixSheet.cell(row, attrEle["col"]).value\
        for attrEle in attrList if attrEle["col"] != 255}

    transEnd = []
    receiveEnd = []
    trLoop = 0
    endStart = getAttrColumn(matrixSheet, "Msg Delay Time") + 1
    while(matrixSheet.cell(1, endStart + trLoop).value):
        if matrixSheet.cell(row, endStart + trLoop).value == "r":
            receiveEnd.append(matrixSheet.cell(1, endStart + trLoop).value)
        elif matrixSheet.cell(row, endStart + trLoop).value == "s":
            transEnd.append(matrixSheet.cell(1, endStart + trLoop).value)

        trLoop += 1
    msg['transBy'] = transEnd
    msg['receiveBy'] = receiveEnd

    return msg

def getSignalInfo(matrixSheet, row):
    attrList = [
        {"item": "name",        "col": getAttrColumn(matrixSheet, "Signal Name")},
        {"item": "desc",        "col": getAttrColumn(matrixSheet, "Signal Description")},
        {"item": "orderType",   "col": getAttrColumn(matrixSheet, "Byte Order")},
        {"item": "startBit",    "col": getAttrColumn(matrixSheet, "Start Bit")},
        {"item": "sendType",    "col": getAttrColumn(matrixSheet, "Signal Send Type")},
        {"item": "len",         "col": getAttrColumn(matrixSheet, "Bit Length")},
        {"item": "dataType",    "col": getAttrColumn(matrixSheet, "Date Type")},
        {"item": "resolution",  "col": getAttrColumn(matrixSheet, "Resolution")},
        {"item": "offset",      "col": getAttrColumn(matrixSheet, "Offset")},
        {"item": "phyMin",      "col": getAttrColumn(matrixSheet, "Signal Min. Value.*(phys)")},
        {"item": "phyMax",      "col": getAttrColumn(matrixSheet, "Signal Max. Value.*(phys)")},
        {"item": "initVal",     "col": getAttrColumn(matrixSheet, "Initial Value")},
        {"item": "unit",        "col": getAttrColumn(matrixSheet, "Unit")},
        {"item": "valDesc",     "col": getAttrColumn(matrixSheet, "Signal Value Description")},
    ]
    signalList = []
    signalLoop = 0
    msgId = matrixSheet.cell(row, getAttrColumn(matrixSheet, "Msg ID")).value
    row += 1

    while(matrixSheet.cell(row + signalLoop,\
        [attrEle["col"] for attrEle in attrList if attrEle.get("item") == "name"][0]).value):
        signal = {attrEle["item"] : (matrixSheet.cell(row + signalLoop, attrEle["col"]).value\
                if matrixSheet.cell(row + signalLoop, attrEle["col"]).value else "")\
            for attrEle in attrList if attrEle["col"] != 255}

        transEnd = []
        receiveEnd = []
        trLoop = 0
        endStart = getAttrColumn(matrixSheet, "Msg Delay Time") + 1
        while(matrixSheet.cell(1, endStart + trLoop).value):
            if matrixSheet.cell(row + signalLoop, endStart + trLoop).value == "r":
                receiveEnd.append(matrixSheet.cell(1, endStart + trLoop).value)
            elif matrixSheet.cell(row + signalLoop, endStart + trLoop).value == "s":
                transEnd.append(matrixSheet.cell(1, endStart + trLoop).value)

            trLoop += 1
        signal['transBy'] = transEnd
        signal['receiveBy'] = receiveEnd
        signal['msgId'] = msgId

        signalList.append(signal)
        signalLoop += 1

    return signalList

def genMatrixBOMsg(msg):
    BOMsg = "BO_" + " "
    BOMsg += str(int(msg['id'], 16)) + " "
    BOMsg += msg['name'] + ":" + " "
    BOMsg += msg['len']

    if len(msg['transBy']):
        for txEnd in msg['transBy']:
            BOMsg += " " + txEnd
    else:
        BOMsg += " " + "Vector__XXX"

    BOMsg += "\n"

    return BOMsg

def genMatrixBOSignal(signalList):
    BOSignals = ""

    for signalEle in signalList:
        BOSignal = " "
        BOSignal += "SG_" + " "
        BOSignal += signalEle['name'] + " : "

        BOSignal += str(int(signalEle['startBit']) + int(signalEle['len']) - 1)\
            if int(signalEle['startBit']) % 8 + int(signalEle['len']) - 1 < 8\
            else str(\
                (int(signalEle['startBit']) + (8 - int(signalEle['startBit']) % 8)) - int(signalEle['len']) - \
                (((int(signalEle['startBit']) + (8 - int(signalEle['startBit']) % 8)) - int(signalEle['len'])) % 8) + \
                (int(signalEle['len']) - 1) % 8)

        BOSignal += "|" + signalEle['len']
        BOSignal += "@"
        BOSignal += "0" if signalEle['orderType'] == "Motorola LSB" else "1"
        BOSignal += "+" if signalEle['dataType'] == "unsigned" or "Unsigned" else "-"
        BOSignal += " "

        BOSignal += "(" + signalEle['resolution'] + "," + signalEle['offset'] + ")" + " "
        BOSignal += "[" + signalEle['phyMin'] + "|" + signalEle['phyMax'] + "]" + " "

        if signalEle['unit']:
            BOSignal += "\"" + signalEle['unit'] + "\""
        else:
            BOSignal += "\"" + "\""
        BOSignal += "  "

        if len(signalEle['receiveBy']):
            for receiveEnd in signalEle['receiveBy']:
                if receiveEnd != signalEle['receiveBy'][0]:
                    BOSignal += ","
                BOSignal += receiveEnd
        else:
            BOSignal += "Vector__XXX"

        BOSignal +="\n"
        BOSignals += BOSignal

    return BOSignals

def genMatrixBO(msg, signalList):
    return genMatrixBOMsg(msg) + genMatrixBOSignal(signalList)

def getMatrixBOS(matrixFile):
    matrixSheet = getMatrixSheet(matrixFile)
    BOS = ""

    for row in range(2, matrixSheet.max_row):
        if matrixSheet.cell(row, 1).value:
            msg = getMsgInfo(matrixSheet, row)
            signalList = getSignalInfo(matrixSheet, row)
            BO = genMatrixBO(msg, signalList)
            BOS += BO + "\n"

    return BOS

def genMatrixCMHeader():
    CMHeader = ""

    CMHeader += "CM_ \""
    CMHeader += getDetailVersion()
    CMHeader += "\";\n"

    return CMHeader

def genMatrixCMMsg(msg):
    CMMsg = ""

    CMMsg += "CM_" + " " + "BO_" + " "
    CMMsg += str(int(msg['id'], 16)) + " "
    CMMsg += "\"" + (msg['desc'] if msg['desc'] else "")  + "\"" + ";"
    CMMsg += "\n"

    return CMMsg

def genMatrixCMSignal(signalList, msgId):
    CMSignals = ""

    for signalEle in signalList:
        CMSignal = "CM_" + " " + "SG_" + " " + str(int(msgId, 16)) + " "
        CMSignal += signalEle['name'] + " "
        CMSignal += "\"" + signalEle['desc'] + "\"" + ";"

        CMSignal +="\n"
        CMSignals += CMSignal

    return CMSignals

def genMatrixCM(msg, signalList):
    return genMatrixCMMsg(msg) + genMatrixCMSignal(signalList, msg['id'])

def getMatrixCMS(matrixFile):
    matrixSheet = getMatrixSheet(matrixFile)
    CMS = genMatrixCMHeader()

    for row in range(2, matrixSheet.max_row):
        if matrixSheet.cell(row, 1).value:
            msg = getMsgInfo(matrixSheet, row)
            signalList = getSignalInfo(matrixSheet, row)
            CM = genMatrixCM(msg, signalList)
            CMS += CM

    return CMS

def getMatrixBUSType(matrixSheet):
    for row in range(2, matrixSheet.max_row):
        if matrixSheet.cell(row, 1).value:
            msg = getMsgInfo(matrixSheet, row)
            if 'frameFormat' not in msg.keys():
                return "CAN"
            if msg['frameFormat'] in ["StandardCAN_FD", "CAN_FD"]:
                return "CAN FD"

    return "CAN"

def getAttrMsgIsNMMsg(msg):
    return 1 if msg['type'] == "NM" else 0

def getAttrMsgSendType(msg):
    typeList = ["Cycle", "CA", "CE", "Event", "NotUsed", "NotUsed", "NotUsed", "IfActive", "NoMsgSendType", "NotUsed"]
    return typeList.index(msg['sendType']) if msg['sendType'] in typeList else 0

def getAttrMsgCycleTime(msg):
    return msg['cycleTime'] if msg['sendType'] in ["Cycle", "CA", "CE"] else 0

def getAttrMsgFastCycleTime(msg):
    return msg['fastCycleTime'] if msg['sendType'] in ["Cycle", "CA", "CE"] else 0

def getAttrMsgNumOfRepetition(msg):
    return msg['numOfRepetition']

def getAttrMsgDelayTime(msg):
    return msg['delayTime']

def getAttrMsgBaudRateSwitch(msg):
    if 'baudRateSwitch' not in msg.keys():
        return 0
    return msg['baudRateSwitch']

def getAttrMsgFrameFormat(msg):
    frameList = ["StandardCAN", "ExtendedCAN", "reserved", "reserved", "reserved", "reserved", "reserved", "reserved", "reserved", "reserved", "reserved", "reserved", "reserved", "reserved", "StandardCAN_FD", "ExtendedCAN_FD"]
    if 'frameFormat' not in msg.keys():
        return 0
    return frameList.index(msg['frameFormat']) if msg['frameFormat'] in frameList else 0

def getAttrMsgType(msg):
    msgList = ["Normal", "NM", "Diag"]
    return msgList.index(msg['type']) if msg['type'] in msgList else 0

def getAttrSigSendType(signal):
    typeList = ["Cycle", "OnChange", "OnWrite", "IfActive", "OnChangeWithRepetition", "OnWriteWithRepetition", "IfActiveWithRepetition"]
    return typeList.index(signal['sendType']) if signal['sendType'] in typeList else 0

def getAttrSigInitVal(signal):
    return int(signal['initVal'], 16) if signal['initVal'] else 0

def getBASAttr():
    BASAttr = {}

    BASGlobalAttrDefList = [
        {"item": "Manufacturer",            "type": '''STRING''',               "defaultVal": "Weltmeister",            "projectValFun": getOEM},
        {"item": "DBName",                  "type": '''STRING''',               "defaultVal": "CentralComputePlatform", "projectValFun": getProjectName},
        {"item": "VersionYear",             "type": '''INT 0 99''',             "defaultVal": "0",                      "projectValFun": getDBCGenerateYear},
        {"item": "VersionMonth",            "type": '''INT 0 12''',             "defaultVal": "0",                      "projectValFun": getDBCGenerateMonth},
        {"item": "VersionDay",              "type": '''INT 0 31''',             "defaultVal": "0",                      "projectValFun": getDBCGenerateDay},
        {"item": "VersionNumber",           "type": '''INT 0 1000''',           "defaultVal": "0",                      "projectValFun": getToolVersion},
        {"item": "BusType",                 "type": '''STRING''',               "defaultVal": "CAN",                    "projectValFun": getMatrixBUSType},
        {"item": "Baudrate",                "type": '''INT 0 1000000''',        "defaultVal": 500000,                   "projectValFun": ""},
        {"item": "ILTxTimeout",             "type": '''INT 0 65535''',          "defaultVal": 0,                        "projectValFun": ""},
        {"item": "NmType",                  "type": '''STRING''',               "defaultVal": "NmAsr",                  "projectValFun": ""},
        {"item": "NmAsrCanMsgCycleTime",    "type": '''INT 1 65535''',          "defaultVal": 500,                      "projectValFun": ""},
        {"item": "NmAsrMessageCount",       "type": '''INT 1 256''',            "defaultVal": 128,                      "projectValFun": ""},
        {"item": "NmAsrRepeatMessageTime",  "type": '''INT 0 65535''',          "defaultVal": 1500,                     "projectValFun": ""},
        {"item": "NmAsrTimeoutTime",        "type": '''INT 1 65535''',          "defaultVal": 2000,                     "projectValFun": ""},
        {"item": "NmAsrWaitBusSleepTime",   "type": '''INT 0 65535''',          "defaultVal": 2000,                     "projectValFun": ""},
        {"item": "NmAsrBaseAddress",        "type": '''HEX 0 1151''',           "defaultVal": 1024,                     "projectValFun": ""},
    ]
    BASNodeAttrDefList = [
        {"item": "ILUsed",                  "type": '''ENUM "No","Yes"''',      "defaultVal": "Yes",                    "projectValFun": ""},
        {"item": "Channel",                 "type": '''STRING''',               "defaultVal": "",                       "projectValFun": ""},
        {"item": "NodeLayerModules",        "type": '''STRING''',               "defaultVal": "",                       "projectValFun": ""},
        {"item": "NmNode",                  "type": '''ENUM "No","Yes"''',      "defaultVal": "No",                     "projectValFun": ""},
        {"item": "NmAsrCanMsgCycleOffset",  "type": '''INT 0 65535''',          "defaultVal": 0,                        "projectValFun": ""},
        {"item": "NmAsrCanMsgReducedTime",  "type": '''INT 1 65535''',          "defaultVal": 320,                      "projectValFun": ""},
        {"item": "NmStationAddress",        "type": '''HEX 0 255''',            "defaultVal": 0,                        "projectValFun": ""},
        {"item": "NmAsrNodeIdentifier",     "type": '''HEX 0 127''',            "defaultVal": 0,                        "projectValFun": ""},
        {"item": "NmAsrNode",               "type": '''ENUM "No","Yes"''',      "defaultVal": "Yes",                    "projectValFun": ""},
    ]
    BASMsgAttrDefList = [
        {"item": "MsgType",                 "type": '''ENUM "Normal","NM","Diag"''', "defaultVal": "Normal",            "projectValFun": getAttrMsgType},
        {"item": "VFrameFormat",            "type": '''ENUM "StandardCAN","ExtendedCAN","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","StandardCAN_FD","ExtendedCAN_FD"''', "defaultVal": "StandardCAN", "projectValFun": getAttrMsgFrameFormat},
        {"item": "NmMessage",               "type": '''ENUM "No","Yes"''',      "defaultVal": "No",                     "projectValFun": getAttrMsgIsNMMsg},
        {"item": "GenMsgSendType",          "type": '''ENUM "Cycle","CA","CE","Event","NotUsed","NotUsed","NotUsed","IfActive","NoMsgSendType","NotUsed"''', "defaultVal": "Cycle", "projectValFun": getAttrMsgSendType},
        {"item": "GenMsgCycleTime",         "type": '''INT 0 0''',              "defaultVal": 0,                        "projectValFun": getAttrMsgCycleTime},
        {"item": "GenMsgDelayTime",         "type": '''INT 0 0''',              "defaultVal": 0,                        "projectValFun": getAttrMsgDelayTime},
        {"item": "GenMsgCycleTimeFast",     "type": '''INT 0 0''',              "defaultVal": 0,                        "projectValFun": getAttrMsgFastCycleTime},
        {"item": "GenMsgNrOfRepetition",    "type": '''INT 0 0''',              "defaultVal": 0,                        "projectValFun": getAttrMsgNumOfRepetition},
        {"item": "CANFD_BRS",               "type": '''ENUM "0","1"''',         "defaultVal": "0",                      "projectValFun": getAttrMsgBaudRateSwitch},
        {"item": "DiagState",               "type": '''ENUM "No","Yes"''',      "defaultVal": "No",                     "projectValFun": ""},
        {"item": "DiagRequest",             "type": '''ENUM "No","Yes"''',      "defaultVal": "No",                     "projectValFun": ""},
        {"item": "DiagResponse",            "type": '''ENUM "No","Yes"''',      "defaultVal": "No",                     "projectValFun": ""},
        {"item": "DiagFdOnly",              "type": '''ENUM "No","Yes"''',      "defaultVal": "No",                     "projectValFun": ""},
        {"item": "DiagConnection",          "type": '''INT 0 65535''',         "defaultVal": 0,                         "projectValFun": ""},
        {"item": "DiagUudtResponse",        "type": '''ENUM "false","true"''',  "defaultVal": "false",                  "projectValFun": ""},
        {"item": "AppMessage",              "type": '''ENUM "No","Yes"''',      "defaultVal": "No",                     "projectValFun": ""},
        {"item": "GenMsgFastOnStart",       "type": '''INT 0 65535''',         "defaultVal": 0,                         "projectValFun": ""},
        {"item": "GenMsgStartDelayTime",    "type": '''INT 0 100000''',        "defaultVal": 0,                         "projectValFun": ""},
        {"item": "GenMsgILSupport",         "type": '''ENUM "No","Yes"''',      "defaultVal": "No",                     "projectValFun": ""},
        {"item": "TpTxIndex",               "type": '''INT 0 255''',           "defaultVal": 0,                         "projectValFun": ""},
        {"item": "NmAsrMessage",            "type": '''ENUM "No","Yes"''',      "defaultVal": "No",                     "projectValFun": ""},
        {"item": "TpApplType",              "type": '''STRING''',               "defaultVal": "",                       "projectValFun": ""},
    ]
    BASSigAttrDefList = [
        {"item": "GenSigSendType",          "type": '''ENUM "Cycle","OnChange","OnWrite","IfActive","OnChangeWithRepetition","OnWriteWithRepetition","IfActiveWithRepetition"''', "defaultVal": "Cycle", "projectValFun": getAttrSigSendType},
        {"item": "GenSigStartValue",        "type": '''INT 0 0''',              "defaultVal": 0,                        "projectValFun": getAttrSigInitVal},
        {"item": "GenSigInactiveValue",     "type": '''INT 0 0''',              "defaultVal": 0,                        "projectValFun": ""},
        {"item": "GenSigInvalidValue",      "type": '''INT 0 0''',              "defaultVal": 0,                        "projectValFun": ""},
        {"item": "GenSigTimeoutValue",      "type": '''INT 0 65535''',          "defaultVal": 0,                        "projectValFun": ""},
    ]

    BASAttr['global'] = BASGlobalAttrDefList
    BASAttr['node'] = BASNodeAttrDefList
    BASAttr['msg'] = BASMsgAttrDefList
    BASAttr['sig'] = BASSigAttrDefList

    return BASAttr

def genMatrixBASGlobalDefinition(BASGlobalAttrDefList):
    BASGlobalDefinition = "".join(["BA_DEF_" + " \"" +\
        BASGlobalAttrDefEle['item'] + "\" " + \
        BASGlobalAttrDefEle['type'] + ";\n"\
        for BASGlobalAttrDefEle in BASGlobalAttrDefList])

    return BASGlobalDefinition

def genMatrixBASNodeDefinition(BASNodeAttrDefList):
    BASNodeDefinition = "".join(["BA_DEF_" + " " + "BU_" + " \"" +\
        BASNodeAttrDefEle['item'] + "\" " + \
        BASNodeAttrDefEle['type'] + ";\n"\
        for BASNodeAttrDefEle in BASNodeAttrDefList])

    return BASNodeDefinition

def genMatrixBASMsgDefinition(BASMsgAttrDefList):
    BASMsgDefinition = "".join(["BA_DEF_" + " " + "BO_" + " \"" +\
        BASMsgAttrDefEle['item'] + "\" " + \
        BASMsgAttrDefEle['type'] + ";\n"\
        for BASMsgAttrDefEle in BASMsgAttrDefList])

    return BASMsgDefinition

def genMatrixBASSigDefinition(BASSigAttrDefList):
    BASSigDefinition = "".join(["BA_DEF_" + " " + "SG_" + " \"" +\
        BASSigAttrDefEle['item'] + "\" " + \
        BASSigAttrDefEle['type'] + ";\n"\
        for BASSigAttrDefEle in BASSigAttrDefList])

    return BASSigDefinition

def genMatrixBASDefinition(BASAttr):
    return genMatrixBASGlobalDefinition(BASAttr['global']) +\
        genMatrixBASNodeDefinition(BASAttr['node']) +\
        genMatrixBASMsgDefinition(BASAttr['msg']) +\
        genMatrixBASSigDefinition(BASAttr['sig'])

def genMatrixBASGlobalDefaultVal(BASGlobalAttrDefList):
    BASGlobalDefaultVal = ""

    for BASGlobalAttrDefEle in BASGlobalAttrDefList:
        BASGlobalDefaultVal += "BA_DEF_DEF_" + " \"" + BASGlobalAttrDefEle['item'] + "\" "
        if not any(type in BASGlobalAttrDefEle['type'] for type in ["INT", "HEX"]):
            BASGlobalDefaultVal += "\"" + BASGlobalAttrDefEle['defaultVal'] + "\";\n"
        else:
            BASGlobalDefaultVal += str(BASGlobalAttrDefEle['defaultVal']) + ";\n"

    return BASGlobalDefaultVal

def genMatrixBASNodeDefaultVal(BASNodeAttrDefList):
    BASNodeDefaultVal = ""

    for BASNodeAttrDefEle in BASNodeAttrDefList:
        BASNodeDefaultVal += "BA_DEF_DEF_" + " \"" + BASNodeAttrDefEle['item'] + "\" "
        if not any(type in BASNodeAttrDefEle['type'] for type in ["INT", "HEX"]):
            BASNodeDefaultVal += "\"" + BASNodeAttrDefEle['defaultVal'] + "\";\n"
        else:
            BASNodeDefaultVal += str(BASNodeAttrDefEle['defaultVal']) + ";\n"

    return BASNodeDefaultVal

def genMatrixBASMsgDefaultVal(BASMsgAttrDefList):
    BASMsgDefaultVal = ""

    for BASMsgAttrDefEle in BASMsgAttrDefList:
        BASMsgDefaultVal += "BA_DEF_DEF_" + " \"" + BASMsgAttrDefEle['item'] + "\" "
        if not any(type in BASMsgAttrDefEle['type'] for type in ["INT", "HEX"]):
            BASMsgDefaultVal += "\"" + BASMsgAttrDefEle['defaultVal'] + "\";\n"
        else:
            BASMsgDefaultVal += str(BASMsgAttrDefEle['defaultVal']) + ";\n"

    return BASMsgDefaultVal

def genMatrixBASSigDefaultVal(BASSigAttrDefList):
    BASSigDefaultVal = ""

    for BASSigAttrDefEle in BASSigAttrDefList:
        BASSigDefaultVal += "BA_DEF_DEF_" + " \"" + BASSigAttrDefEle['item'] + "\" "
        if not any(type in BASSigAttrDefEle['type'] for type in ["INT", "HEX"]):
            BASSigDefaultVal += "\"" + BASSigAttrDefEle['defaultVal'] + "\";\n"
        else:
            BASSigDefaultVal += str(BASSigAttrDefEle['defaultVal']) + ";\n"

    return BASSigDefaultVal

def genMatrixBASDefaultVal(BASAttr):
    return genMatrixBASGlobalDefaultVal(BASAttr['global']) +\
        genMatrixBASNodeDefaultVal(BASAttr['node']) +\
        genMatrixBASMsgDefaultVal(BASAttr['msg']) +\
        genMatrixBASSigDefaultVal(BASAttr['sig'])

def genMatrixBASGlobalProjectVal(BASGlobalAttrDefList, matrixSheet):
    BASGlobalProjectVal = ""

    for BASGlobalAttrDefEle in BASGlobalAttrDefList:
        if not BASGlobalAttrDefEle['projectValFun']:
            continue

        BASGlobalProjectVal += "BA_" + " \"" + BASGlobalAttrDefEle['item'] + "\" "
        if not any(type in BASGlobalAttrDefEle['type'] for type in ["INT", "HEX"]):
            BASGlobalProjectVal += "\""
            BASGlobalProjectVal += BASGlobalAttrDefEle['projectValFun']() if BASGlobalAttrDefEle['projectValFun'].__code__.co_argcount == 0 else BASGlobalAttrDefEle['projectValFun'](matrixSheet)
            BASGlobalProjectVal += "\";\n"
        else:
            BASGlobalProjectVal += str(BASGlobalAttrDefEle['projectValFun']() if BASGlobalAttrDefEle['projectValFun'].__code__.co_argcount == 0 else BASGlobalAttrDefEle['projectValFun'](matrixSheet)) + ";\n"

    return BASGlobalProjectVal

def genMatrixBASNodeProjectVal(BASNodeAttrDefList, matrixSheet):
    BASNodeProjectVal = ""

    for BASNodeAttrDefEle in BASNodeAttrDefList:
        if not BASNodeAttrDefEle['projectValFun']:
            continue

        BASNodeProjectVal += "BA_" + " \"" + BASNodeAttrDefEle['item'] + "\" "
        if not any(type in BASNodeAttrDefEle['type'] for type in ["INT", "HEX"]):
            BASNodeProjectVal += "\""
            BASNodeProjectVal += BASNodeAttrDefEle['projectValFun']() if BASNodeAttrDefEle['projectValFun'].__code__.co_argcount == 0 else BASNodeAttrDefEle['projectValFun'](matrixSheet)
            BASNodeProjectVal += "\";\n"
        else:
            BASNodeProjectVal += str(BASNodeAttrDefEle['projectValFun']() if BASNodeAttrDefEle['projectValFun'].__code__.co_argcount == 0 else BASNodeAttrDefEle['projectValFun'](matrixSheet)) + ";\n"

    return BASNodeProjectVal

def genMatrixBASMsgProjectVal(BASMsgAttrDefList, matrixSheet):
    BASMsgProjectVal = ""

    for row in range(2, matrixSheet.max_row):
        if matrixSheet.cell(row, 1).value:
            msg = getMsgInfo(matrixSheet, row)

            for BASMsgAttrDefEle in BASMsgAttrDefList:
                if not BASMsgAttrDefEle['projectValFun']:
                    continue

                BASMsgProjectVal += "BA_" + " \"" + BASMsgAttrDefEle['item'] + "\" " + "BO_ "
                BASMsgProjectVal += str(int(msg['id'], 16)) + " "
                if not any(type in BASMsgAttrDefEle['type'] for type in ["INT", "HEX", "ENUM"]):
                    BASMsgProjectVal += "\""
                    BASMsgProjectVal += BASMsgAttrDefEle['projectValFun'](msg)
                    BASMsgProjectVal += "\";\n"
                else:
                    BASMsgProjectVal += str(BASMsgAttrDefEle['projectValFun'](msg)) + ";\n"

    return BASMsgProjectVal

def genMatrixBASSigProjectVal(BASSigAttrDefList, matrixSheet):
    BASSigProjectVal = ""

    for row in range(2, matrixSheet.max_row):
        if matrixSheet.cell(row, 1).value:
            signalList = getSignalInfo(matrixSheet, row)

            for signal in signalList:
                for BASSigAttrDefEle in BASSigAttrDefList:
                    if not BASSigAttrDefEle['projectValFun']:
                        continue

                    BASSigProjectVal += "BA_" + " \"" + BASSigAttrDefEle['item'] + "\" " + "SG_ "
                    BASSigProjectVal += str(int(signal['msgId'], 16)) + " " + signal['name'] + " "
                    if not any(type in BASSigAttrDefEle['type'] for type in ["INT", "HEX", "ENUM"]):
                        BASSigProjectVal += "\""
                        BASSigProjectVal += BASSigAttrDefEle['projectValFun'](signal)
                        BASSigProjectVal += "\";\n"
                    else:
                        BASSigProjectVal += str(BASSigAttrDefEle['projectValFun'](signal)) + ";\n"

    return BASSigProjectVal

def genMatrixBASProjectVal(BASAttr, matrixSheet):
    return genMatrixBASGlobalProjectVal(BASAttr['global'], matrixSheet) +\
        genMatrixBASNodeProjectVal(BASAttr['node'], matrixSheet) +\
        genMatrixBASMsgProjectVal(BASAttr['msg'], matrixSheet) +\
        genMatrixBASSigProjectVal(BASAttr['sig'], matrixSheet)

# 特征名称类型定义:
# BA_DEF_ Object AttributeName ValueType Min Max;
# 特征默认值定义:
# BA_DEF_DEF_ AttributeName DefaultValue;
# 特征项目设置值定义:
# BA_ AttributeName projectValue;
def getMatrixBAS(matrixFile):
    matrixSheet = getMatrixSheet(matrixFile)
    BASAttr = getBASAttr()
    BAS = ""

    BAS += genMatrixBASDefinition(BASAttr)
    BAS += genMatrixBASDefaultVal(BASAttr)
    BAS += genMatrixBASProjectVal(BASAttr, matrixSheet)

    return BAS

def getMatrixSigVAL(signal):
    if not signal['valDesc']:
        return ""

    signalValLoop = 0
    signalVal = ""
    signalValList = [ele.strip() for ele in re.split(r":|：|\n", signal['valDesc']) if ele]

    while(signalValLoop < len(signalValList)):
        if signalValLoop % 2 == 0:
            if "~" in signalValList[signalValLoop]:
                startVal = int(signalValList[signalValLoop].split("~")[0], 16)
                endVal = int(signalValList[signalValLoop].split("~")[1], 16)

                for valLoop in range(startVal, endVal + 1):
                    signalVal += str(valLoop) + " " + "\"" + signalValList[signalValLoop + 1] + "\""
                    signalVal += " " if valLoop != endVal else ""

                signalValLoop += 2
                continue
            signalVal += str(int(signalValList[signalValLoop], 16))
        else:
            signalVal += "\"" + signalValList[signalValLoop] + "\""

        signalVal += " " if signalValLoop != len(signalValList) - 1 else ""
        signalValLoop += 1

    return signalVal

def getMatrixVAL(matrixFile):
    matrixSheet = getMatrixSheet(matrixFile)
    VAL = ""

    for row in range(2, matrixSheet.max_row):
        if matrixSheet.cell(row, 1).value:
            signalList = getSignalInfo(matrixSheet, row)

            for signal in signalList:
                VAL += "VAL_ " + str(int(signal['msgId'], 16)) + " "
                VAL += signal['name'] + " "
                VAL += getMatrixSigVAL(signal) + ";\n"

    return VAL
